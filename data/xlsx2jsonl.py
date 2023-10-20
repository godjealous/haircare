import openpyxl
import json
import os

fout = open('QA.jsonl','w')
workbook = openpyxl.load_workbook('QA.xlsx')
sheet = workbook.active  # 或者使用 sheet = workbook['工作表名称']
for idx, row in enumerate(sheet.iter_rows(values_only=True)):
    if idx == 0: continue
    Q = row[0]
    A = row[1]
    data = {}
    data['Q'] = Q
    data['A'] = A
    fout.write(json.dumps(data, ensure_ascii=False))
    fout.write("\n")
fout.close()



fout = open('images_text.jsonl','w')
workbook = openpyxl.load_workbook('images_text.xlsx')
sheet = workbook.active  # 或者使用 sheet = workbook['工作表名称']
for idx, row in enumerate(sheet.iter_rows(values_only=True)):
    if idx == 0: continue
    data = {}
    data['path'] = 'images/'+row[0]
    if not os.path.exists(data['path']):
        print(data['path'])

    data['consult'] = row[1]
    data['level'] = row[2]
    data['classification'] = row[3]
    data['diagnosis'] = row[4]
    data['treatment'] = row[5]
    data['gotodoctor'] = row[6]
    fout.write(json.dumps(data, ensure_ascii=False))
    fout.write("\n")
fout.close()
